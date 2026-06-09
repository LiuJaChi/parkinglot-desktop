# -*- coding: utf-8 -*-
"""
車輛通行記錄導入管理系統 - 完整版本
支持: 時間、車道、車號、身份、結果、電子標籤
查詢條件: 車號(輸入)、身份(下拉)、結果(下拉)、日期範圍
車號為空白/NAN/nan/null 的記錄 - 完全忽略不顯示
支持匯入格式: Excel (.xlsx, .xls) 和 CSV (.csv)
支持輸出格式: PDF 報表
統計方式: 身份統計按不重複車輛數量計算

依賴庫：openpyxl, pillow, reportlab
"""

import json
import os
import subprocess
import platform
import sys
import csv
import queue
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Set
from dataclasses import dataclass, asdict
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from PIL import Image, ImageTk
import tempfile
import calendar
from collections import defaultdict

# ============================================================================
# 依賴庫檢查
# ============================================================================

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    messagebox.showerror("錯誤", "缺少 openpyxl 庫\n請執行: pip install openpyxl")
    sys.exit(1)

# ReportLab 檢查（用於 PDF 生成）
REPORTLAB_AVAILABLE = False
FONT_REGISTERED = False
FONT_NAME = 'Helvetica'

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    pass


# ============================================================================
# 中文字體註冊
# ============================================================================

def register_chinese_fonts():
    """註冊中文字體支持"""
    global FONT_REGISTERED, FONT_NAME
    
    if not REPORTLAB_AVAILABLE:
        return False
    
    try:
        font_paths = []
        
        if sys.platform == 'win32':
            font_paths = [
                r"C:\Windows\Fonts\simsun.ttc",
                r"C:\Windows\Fonts\simhei.ttf",
                r"C:\Windows\Fonts\msyh.ttf",
                r"C:\Windows\Fonts\msyh.ttc",
            ]
        elif sys.platform == 'darwin':
            font_paths = [
                "/System/Library/Fonts/PingFang.ttc",
                "/Library/Fonts/SimSun.ttf",
                "/System/Library/Fonts/Songti.ttc",
            ]
        else:
            font_paths = [
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/droid/DroidSansFallback.ttf",
            ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    if font_path.endswith('.ttf'):
                        pdfmetrics.registerFont(TTFont('SimSun', font_path))
                    elif font_path.endswith('.ttc'):
                        try:
                            pdfmetrics.registerFont(TTFont('SimSun', font_path))
                        except:
                            continue
                    
                    FONT_REGISTERED = True
                    FONT_NAME = 'SimSun'
                    return True
                except:
                    continue
        
        FONT_NAME = 'Helvetica'
        return False
        
    except Exception:
        FONT_NAME = 'Helvetica'
        return False


if REPORTLAB_AVAILABLE:
    register_chinese_fonts()


# ============================================================================
# 數據模型
# ============================================================================

@dataclass
class VehicleRecord:
    """車輛通行記錄"""
    id: int
    time: str
    lane: str
    plate_number: str
    identity: str
    result: str
    etag: str


# ============================================================================
# 進度窗口
# ============================================================================

class ProgressWindow:
    """進度顯示窗口"""

    def __init__(self, parent, title="導入進度"):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("400x150")
        self.window.resizable(False, False)
        self.window.transient(parent)
        self.window.grab_set()

        self.is_closed = False
        self.progress_queue = queue.Queue()

        tk.Label(self.window, text=title, font=("微軟雅黑", 12, "bold"), bg="white", fg="#333").pack(pady=10)

        progress_frame = tk.Frame(self.window, bg="white")
        progress_frame.pack(pady=10, padx=20, fill=tk.X)

        self.progress = ttk.Progressbar(progress_frame, length=350, mode='determinate', maximum=100)
        self.progress.pack(fill=tk.X)

        self.percent_label = tk.Label(self.window, text="0%", font=("微軟雅黑", 14, "bold"), bg="white", fg="#2196F3")
        self.percent_label.pack(pady=5)

        self.detail_label = tk.Label(self.window, text="正在處理...", font=("微軟雅黑", 9), bg="white", fg="#666")
        self.detail_label.pack(pady=5)

        self.monitor_queue()
        self.window.update()

    def monitor_queue(self):
        try:
            while True:
                try:
                    current, total, detail_text = self.progress_queue.get_nowait()
                    self.update_progress_ui(current, total, detail_text)
                except queue.Empty:
                    break
        except:
            pass

        if not self.is_closed:
            self.window.after(100, self.monitor_queue)

    def update_progress(self, current: int, total: int, detail_text: str = ""):
        """更新進度（線程安全）"""
        try:
            self.progress_queue.put((current, total, detail_text))
        except:
            pass

    def update_progress_ui(self, current: int, total: int, detail_text: str = ""):
        try:
            if total == 0:
                total = 1
            percentage = int((current / total) * 100)
            self.progress['value'] = percentage
            self.percent_label.config(text="{0}%".format(percentage))
            self.detail_label.config(text=detail_text if detail_text else "已處理: {0}/{1} 筆".format(current, total))
            self.window.update_idletasks()
        except:
            pass

    def close(self):
        """關閉進度窗口"""
        try:
            self.is_closed = True
            self.window.destroy()
        except:
            pass


# ============================================================================
# 日期選擇器
# ============================================================================

class DatePicker(tk.Toplevel):
    """日期選擇器"""

    def __init__(self, parent, initial_date=None):
        super().__init__(parent)
        self.title("選擇日期")
        self.geometry("400x350")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        if initial_date is None:
            initial_date = datetime.now()
        elif isinstance(initial_date, str):
            initial_date = datetime.strptime(initial_date, "%Y-%m-%d")

        self.selected_date = None
        self.current_year = initial_date.year
        self.current_month = initial_date.month
        self.initial_day = initial_date.day

        self._build_ui()
        self.center_window()
        self.wait_window()

    def _build_ui(self):
        """構建 UI"""
        header_frame = tk.Frame(self, bg="#2196F3")
        header_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(header_frame, text="◀", command=self.prev_month, bg="#1976D2", fg="white", cursor="hand2").pack(side=tk.LEFT, padx=2)
        self.date_label = tk.Label(header_frame, text="{0}年 {1}月".format(self.current_year, self.current_month), bg="#2196F3", fg="white", font=("微軟雅黑", 12, "bold"))
        self.date_label.pack(side=tk.LEFT, expand=True)
        tk.Button(header_frame, text="▶", command=self.next_month, bg="#1976D2", fg="white", cursor="hand2").pack(side=tk.RIGHT, padx=2)

        calendar_frame = tk.Frame(self)
        calendar_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        weeks = ["一", "二", "三", "四", "五", "六", "日"]
        for i, week in enumerate(weeks):
            tk.Label(calendar_frame, text=week, bg="#E0E0E0", fg="#333", font=("微軟雅黑", 9, "bold"), width=5, height=2).grid(row=0, column=i, sticky="nsew")

        self.day_buttons = {}
        cal = calendar.monthcalendar(self.current_year, self.current_month)

        for week_num, week in enumerate(cal, start=1):
            for day_num, day in enumerate(week):
                if day == 0:
                    tk.Label(calendar_frame, text="", bg="white").grid(row=week_num, column=day_num, sticky="nsew")
                else:
                    is_initial = (day == self.initial_day)
                    btn = tk.Button(calendar_frame, text=str(day), width=5, height=2, font=("微軟雅黑", 10, "bold" if is_initial else "normal"),
                                   bg="#FFD54F" if is_initial else "#FFFFFF", fg="#333", cursor="hand2", command=lambda d=day: self.select_date(d))
                    btn.grid(row=week_num, column=day_num, sticky="nsew", padx=1, pady=1)
                    self.day_buttons[day] = btn
                calendar_frame.grid_columnconfigure(day_num, weight=1)
            calendar_frame.grid_rowconfigure(week_num, weight=1)

        button_frame = tk.Frame(self)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(button_frame, text="今天", command=self.select_today, bg="#4CAF50", fg="white", font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.LEFT, padx=2)
        tk.Button(button_frame, text="確定", command=self.confirm, bg="#2196F3", fg="white", font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.RIGHT, padx=2)
        tk.Button(button_frame, text="取消", command=self.cancel, bg="#757575", fg="white", font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.RIGHT, padx=2)

    def prev_month(self):
        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        self.refresh_calendar()

    def next_month(self):
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        self.refresh_calendar()

    def refresh_calendar(self):
        self.date_label.config(text="{0}年 {1}月".format(self.current_year, self.current_month))
        for btn in self.day_buttons.values():
            btn.destroy()
        self.day_buttons.clear()

        calendar_frame = self.winfo_children()[1]
        cal = calendar.monthcalendar(self.current_year, self.current_month)

        for week_num, week in enumerate(cal, start=1):
            for day_num, day in enumerate(week):
                if day == 0:
                    continue
                btn = tk.Button(calendar_frame, text=str(day), width=5, height=2, font=("微軟雅黑", 10), bg="#FFFFFF", fg="#333", cursor="hand2", command=lambda d=day: self.select_date(d))
                btn.grid(row=week_num, column=day_num, sticky="nsew", padx=1, pady=1)
                self.day_buttons[day] = btn

    def select_date(self, day):
        self.selected_date = datetime(self.current_year, self.current_month, day)
        self.confirm()

    def select_today(self):
        today = datetime.now()
        self.current_year = today.year
        self.current_month = today.month
        self.selected_date = today
        self.confirm()

    def confirm(self):
        if self.selected_date is None:
            self.selected_date = datetime(self.current_year, self.current_month, 1)
        self.destroy()

    def cancel(self):
        self.selected_date = None
        self.destroy()

    def center_window(self):
        self.update_idletasks()
        x = self.master.winfo_x() + (self.master.winfo_width() // 2) - (self.winfo_width() // 2)
        y = self.master.winfo_y() + (self.master.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry("+{0}+{1}".format(x, y))


# ============================================================================
# 業務邏輯層
# ============================================================================

class VehicleImporter:
    """車輛記錄導入器"""

    def __init__(self, json_file: str = "vehicle_data.json"):
        self.json_file = json_file
        self.records: List[VehicleRecord] = []
        self.photo_directory = ""
        self.vehicle_image_directory = ""
        self.progress_callback = None
        self.load_data()

    def set_photo_directory(self, directory: str):
        self.photo_directory = directory

    def set_vehicle_image_directory(self, directory: str):
        self.vehicle_image_directory = directory

    def set_progress_callback(self, callback):
        self.progress_callback = callback

    def load_data(self):
        if os.path.exists(self.json_file):
            try:
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.records = [VehicleRecord(**record) for record in data]
                return
            except:
                pass
        self._init_sample_data()

    def _init_sample_data(self):
        self.records = [
            VehicleRecord(1, "2024-01-15 08:30:00", "車道1", "粵B12345", "業主", "入場", "TAG001"),
            VehicleRecord(2, "2024-01-15 08:35:00", "車道2", "粵B54321", "訪客", "入場", "TAG002"),
            VehicleRecord(3, "2024-01-15 08:40:00", "車道1", "粵B12345", "業主", "無法入場", "TAG001"),
            VehicleRecord(4, "2024-01-15 09:00:00", "車道3", "", "員工", "無法入場", "TAG003"),
            VehicleRecord(5, "2024-01-15 09:15:00", "車道2", "粵C11111", "訪客", "預警", "TAG004"),
            VehicleRecord(6, "2024-01-16 10:00:00", "車道1", "粵B12345", "業主", "入場", "TAG001"),
            VehicleRecord(7, "2024-01-16 10:30:00", "車道2", "", "訪客", "無法入場", "TAG005"),
            VehicleRecord(8, "2024-01-17 14:20:00", "車道3", "粵D99999", "員工", "無法入場", "TAG006"),
            VehicleRecord(9, "2024-01-17 15:00:00", "車道1", "粵E55555", "訪客", "預警", "TAG007"),
            VehicleRecord(10, "2024-01-17 15:30:00", "車道2", "粵F66666", "業主", "入場", "TAG008"),
        ]
        self.save_data()

    def save_data(self):
        try:
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump([asdict(record) for record in self.records], f, ensure_ascii=False, indent=2)
        except:
            pass

    @staticmethod
    def normalize_plate_number(plate_number: Any) -> str:
        if plate_number is None:
            return ""
        plate_str = str(plate_number).strip()
        if not plate_str or plate_str.lower() in ["null", "nan"]:
            return ""
        return plate_str

    @staticmethod
    def is_valid_record(plate_number: str) -> bool:
        return bool(plate_number)

    @staticmethod
    def validate_format(headers: List[str]) -> Tuple[bool, str]:
        required = ['時間', '車道', '車號', '身份', '結果', '電子標籤']
        missing = [col for col in required if col not in headers]
        if missing:
            return False, "缺少列: {0}".format(', '.join(missing))
        return True, "格式正確"

    def import_from_excel(self, excel_file: str, append_mode: bool = True) -> Tuple[bool, str, int]:
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]

            is_valid, message = self.validate_format(headers)
            if not is_valid:
                return False, message, 0

            col_indices = {header: headers.index(header) for header in headers}

            if not append_mode:
                self.records = []

            next_id = max([r.id for r in self.records], default=0) + 1
            imported_count = 0
            total_rows = worksheet.max_row - 1

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if self.progress_callback:
                        self.progress_callback(row_idx - 1, total_rows, "正在匯入...")

                    time = str(row[col_indices['時間']]).strip() if row[col_indices['時間']] else ""
                    lane = str(row[col_indices['車道']]).strip() if row[col_indices['車道']] else ""
                    plate_number = self.normalize_plate_number(row[col_indices['車號']])
                    identity = str(row[col_indices['身份']]).strip() if row[col_indices['身份']] else ""
                    result = str(row[col_indices['結果']]).strip() if row[col_indices['結果']] else ""
                    etag = str(row[col_indices['電子標籤']]).strip() if row[col_indices['電子標籤']] else ""

                    if not plate_number or not all([time, lane, identity, result, etag]):
                        continue

                    self.records.append(VehicleRecord(next_id, time, lane, plate_number, identity, result, etag))
                    next_id += 1
                    imported_count += 1

                except:
                    pass

            workbook.close()
            self.save_data()
            return True, "成功匯入 {0} 筆記錄".format(imported_count), imported_count

        except Exception as e:
            return False, "匯入失敗: {0}".format(str(e)), 0

    def import_from_csv(self, csv_file: str, append_mode: bool = True) -> Tuple[bool, str, int]:
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers_row = next(reader, None)
                if not headers_row:
                    return False, "CSV 格式錯誤", 0

                headers = [h.strip() for h in headers_row]
                is_valid, message = self.validate_format(headers)
                if not is_valid:
                    return False, message, 0

                col_indices = {header: headers.index(header) for header in headers}

                if not append_mode:
                    self.records = []

                next_id = max([r.id for r in self.records], default=0) + 1
                imported_count = 0
                all_rows = list(reader)

                for row_idx, row in enumerate(all_rows, start=2):
                    try:
                        if self.progress_callback:
                            self.progress_callback(row_idx - 1, len(all_rows), "正在匯入...")

                        while len(row) <= max(col_indices.values()):
                            row.append('')

                        time = row[col_indices['時間']].strip() if col_indices['時間'] < len(row) else ""
                        lane = row[col_indices['車道']].strip() if col_indices['車道'] < len(row) else ""
                        plate_number = self.normalize_plate_number(row[col_indices['車號']].strip() if col_indices['車號'] < len(row) else "")
                        identity = row[col_indices['身份']].strip() if col_indices['身份'] < len(row) else ""
                        result = row[col_indices['結果']].strip() if col_indices['結果'] < len(row) else ""
                        etag = row[col_indices['電子標籤']].strip() if col_indices['電子標籤'] < len(row) else ""

                        if not plate_number or not all([time, lane, identity, result, etag]):
                            continue

                        self.records.append(VehicleRecord(next_id, time, lane, plate_number, identity, result, etag))
                        next_id += 1
                        imported_count += 1

                    except:
                        pass

            self.save_data()
            return True, "成功匯入 {0} 筆記錄".format(imported_count), imported_count

        except Exception as e:
            return False, "匯入失敗: {0}".format(str(e)), 0

    def import_from_file(self, file_path: str, append_mode: bool = True) -> Tuple[bool, str, int]:
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.xlsx', '.xls']:
            return self.import_from_excel(file_path, append_mode)
        elif file_ext == '.csv':
            return self.import_from_csv(file_path, append_mode)
        else:
            return False, "不支持的格式", 0

    def get_all_records(self) -> List[Dict]:
        return [asdict(record) for record in self.records if self.is_valid_record(record.plate_number)]

    def get_unique_values(self) -> Dict[str, List[str]]:
        valid_records = [r for r in self.records if self.is_valid_record(r.plate_number)]
        return {
            "identities": sorted(set(r.identity for r in valid_records)),
            "results": sorted(set(r.result for r in valid_records)),
        }

    def get_date_range(self) -> Tuple[str, str]:
        valid_records = [r for r in self.records if self.is_valid_record(r.plate_number)]
        if not valid_records:
            return "", ""
        dates = [r.time.split()[0] for r in valid_records if r.time]
        return (min(dates), max(dates)) if dates else ("", "")

    def search_by_conditions(self, plate_number: str = "", identity: str = "", result: str = "", start_date: str = "", end_date: str = "") -> List[Dict]:
        results = []
        for record in self.records:
            if not self.is_valid_record(record.plate_number):
                continue
            if plate_number and plate_number not in record.plate_number:
                continue
            if identity and identity != record.identity:
                continue
            if result and result != record.result:
                continue
            if start_date or end_date:
                try:
                    record_date = record.time.split()[0]
                    if start_date and record_date < start_date:
                        continue
                    if end_date and record_date > end_date:
                        continue
                except:
                    continue
            results.append(asdict(record))
        return results

    def get_statistics(self, start_date: str = "", end_date: str = "") -> Dict[str, Any]:
        """✅ 修正: 身份統計按不重複車輛數量計算"""
        records_to_analyze = []
        if start_date or end_date:
            for record in self.records:
                if not self.is_valid_record(record.plate_number):
                    continue
                try:
                    record_date = record.time.split()[0]
                    if start_date and record_date < start_date:
                        continue
                    if end_date and record_date > end_date:
                        continue
                    records_to_analyze.append(record)
                except:
                    continue
        else:
            records_to_analyze = [r for r in self.records if self.is_valid_record(r.plate_number)]

        if not records_to_analyze:
            return {
                "total": 0, "no_entry_total": 0, "entry_total": 0, "warning_total": 0,
                "no_entry_vehicles": {}, "no_entry_identities": {}, "no_entry_last_time": {},
                "entry_vehicles": {}, "entry_identities": {},
                "warning_identities": {},
                "date_range": (start_date, end_date)
            }

        # ✅ 按車牌統計 (相同車牌同天只計一次)
        no_entry_vehicles_per_day: Set[Tuple[str, str]] = set()  # (date, plate)
        entry_vehicles_per_day: Set[Tuple[str, str]] = set()
        warning_vehicles_per_day: Set[Tuple[str, str]] = set()
        
        # ✅ 按身份和結果統計 (用於后續計算不重複的車輛)
        no_entry_by_identity: Dict[str, Set[str]] = defaultdict(set)  # identity -> set of plates
        entry_by_identity: Dict[str, Set[str]] = defaultdict(set)
        warning_by_identity: Dict[str, Set[str]] = defaultdict(set)
        
        # ✅ 記錄最後無法入場的時間
        no_entry_last_time: Dict[str, str] = {}

        for record in records_to_analyze:
            try:
                record_date = record.time.split()[0]
                
                if record.result == "無法入場":
                    no_entry_vehicles_per_day.add((record_date, record.plate_number))
                    no_entry_by_identity[record.identity].add(record.plate_number)
                    
                    # 保留最新的無法入場時間
                    if record.plate_number not in no_entry_last_time or record.time > no_entry_last_time[record.plate_number]:
                        no_entry_last_time[record.plate_number] = record.time
                        
                elif record.result == "入場":
                    entry_vehicles_per_day.add((record_date, record.plate_number))
                    entry_by_identity[record.identity].add(record.plate_number)
                        
                elif record.result == "預警":
                    warning_vehicles_per_day.add((record_date, record.plate_number))
                    warning_by_identity[record.identity].add(record.plate_number)
            except:
                pass

        # ✅ 統計各個身份的不重複車輛數
        no_entry_identities = {identity: len(plates) for identity, plates in no_entry_by_identity.items()}
        entry_identities = {identity: len(plates) for identity, plates in entry_by_identity.items()}
        warning_identities = {identity: len(plates) for identity, plates in warning_by_identity.items()}

        # ✅ 統計各個車牌的天數
        no_entry_vehicles = defaultdict(int)
        entry_vehicles = defaultdict(int)
        
        for (date, plate) in no_entry_vehicles_per_day:
            no_entry_vehicles[plate] += 1
        for (date, plate) in entry_vehicles_per_day:
            entry_vehicles[plate] += 1

        # ✅ 統計總數：計算不重複的車牌數
        all_no_entry_plates = set(plate for (date, plate) in no_entry_vehicles_per_day)
        all_entry_plates = set(plate for (date, plate) in entry_vehicles_per_day)
        all_warning_plates = set(plate for (date, plate) in warning_vehicles_per_day)
        
        total_plates = all_no_entry_plates | all_entry_plates | all_warning_plates

        stats = {
            "total": len(total_plates),
            "no_entry_total": len(all_no_entry_plates),
            "entry_total": len(all_entry_plates),
            "warning_total": len(all_warning_plates),
            "no_entry_vehicles": no_entry_vehicles,
            "no_entry_identities": no_entry_identities,  # ✅ 按不重複車輛數
            "no_entry_last_time": no_entry_last_time,
            "entry_vehicles": entry_vehicles,
            "entry_identities": entry_identities,  # ✅ 按不重複車輛數
            "warning_identities": warning_identities,  # ✅ 按不重複車輛數
            "date_range": (start_date, end_date) if (start_date or end_date) else self.get_date_range()
        }

        return stats

    def find_vehicle_photo(self, plate_number: str, datetime_str: str) -> str:
        """查找車輛照片"""
        if not self.photo_directory or not plate_number:
            return ""
        try:
            date_part = datetime_str.split()[0]
            date_yyyymmdd = date_part.replace("-", "")
            date_dir = os.path.join(self.photo_directory, date_yyyymmdd)
            
            if os.path.exists(date_dir):
                for filename in os.listdir(date_dir):
                    if (filename.lower().endswith((".jpg", ".jpeg", ".png")) and
                            plate_number in filename and "入口" in filename):
                        return os.path.join(date_dir, filename)
            return ""
        except:
            return ""

    def find_vehicle_image(self, plate_number: str, datetime_str: str) -> str:
        """查找車輛影像（視頻）"""
        if not self.vehicle_image_directory or not plate_number:
            return ""
        try:
            date_part = datetime_str.split()[0]
            date_yyyymmdd = date_part.replace("-", "")
            time_part = datetime_str.split()[1]
            time_hhmmss = time_part.replace(":", "")
            target_time = int(time_hhmmss)

            date_dir = os.path.join(self.vehicle_image_directory, date_yyyymmdd)
            if not os.path.exists(date_dir):
                return ""

            matching_files = []
            for filename in os.listdir(date_dir):
                if filename.lower().endswith(".mp4") and "入口" in filename:
                    try:
                        parts = filename.split('_')
                        if len(parts) > 0:
                            time_part_str = parts[0]
                            if '-' in time_part_str:
                                file_time_str = time_part_str.split('-')[1]
                                file_time = int(file_time_str)
                                if file_time == target_time:
                                    return os.path.join(date_dir, filename)
                                if file_time < target_time:
                                    matching_files.append((file_time, filename))
                    except:
                        continue

            if matching_files:
                matching_files.sort(key=lambda x: x[0], reverse=True)
                closest_file = matching_files[0][1]
                return os.path.join(date_dir, closest_file)

            return ""
        except:
            return ""


# ============================================================================
# PDF 報表生成器
# ============================================================================

class PDFReportGenerator:
    """PDF 報表生成器"""

    @staticmethod
    def generate_statistics_report(output_path: str, stats: Dict[str, Any], title: str = "Vehicle Traffic Statistics") -> Tuple[bool, str]:
        """生成統計報表 PDF"""
        if not REPORTLAB_AVAILABLE:
            return False, "缺少 reportlab 庫\n請執行: pip install reportlab"

        try:
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    return False, "無法創建輸出目錄: {0}".format(str(e))
            
            test_path = output_path
            if not output_dir:
                test_path = os.path.join(os.getcwd(), output_path)
            
            try:
                test_dir = os.path.dirname(test_path) or os.getcwd()
                if not os.access(test_dir, os.W_OK):
                    return False, "無寫入權限到目錄: {0}".format(test_dir)
            except Exception as e:
                return False, "權限檢查失敗: {0}".format(str(e))
            
            doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()

            # 標題
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=12,
                alignment=TA_CENTER,
                fontName=FONT_NAME
            )
            title_text = "車輛通行統計報表"
            elements.append(Paragraph(title_text, title_style))
            elements.append(Spacer(1, 12))

            # 生成時間
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            time_style = ParagraphStyle('TimeStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, fontName=FONT_NAME)
            elements.append(Paragraph("生成時間: {0}".format(now), time_style))
            elements.append(Spacer(1, 12))

            # 統計摘要表格
            no_entry_percent = round(stats['no_entry_total'] * 100 / stats['total']) if stats['total'] > 0 else 0
            entry_percent = round(stats['entry_total'] * 100 / stats['total']) if stats['total'] > 0 else 0
            warning_percent = round(stats['warning_total'] * 100 / stats['total']) if stats['total'] > 0 else 0

            summary_data = [
                ["統計項目", "車子數量", "百分比"],
                ["總計", str(stats['total']), "100%"],
                ["入場", str(stats['entry_total']), "{0}%".format(entry_percent)],
                ["無法入場", str(stats['no_entry_total']), "{0}%".format(no_entry_percent)],
                ["預警", str(stats['warning_total']), "{0}%".format(warning_percent)],
            ]

            summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            # 車號統計 - 包含最後無法入場時間
            plate_title_style = ParagraphStyle('Heading2', parent=styles['Heading2'], fontName=FONT_NAME)
            elements.append(Paragraph("無法入場車號統計 (含最後無法入場時間)", plate_title_style))
            elements.append(Spacer(1, 10))

            no_entry_vehicles = sorted(stats['no_entry_vehicles'].items(), key=lambda x: x[1], reverse=True)
            if no_entry_vehicles:
                no_entry_data = [["車號", "無法入場天數", "最後無法入場時間"]]
                for plate, count in no_entry_vehicles[:20]:
                    last_time = stats.get('no_entry_last_time', {}).get(plate, "")
                    no_entry_data.append([plate, str(count), last_time])
                
                no_entry_table = Table(no_entry_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
                no_entry_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F44336')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                elements.append(no_entry_table)
            else:
                elements.append(Paragraph("No records", ParagraphStyle('Normal', parent=styles['Normal'], fontName=FONT_NAME)))

            elements.append(Spacer(1, 20))

            # 身份統計 - ✅ 按不重複車輛數量
            identity_title_style = ParagraphStyle('Heading2', parent=styles['Heading2'], fontName=FONT_NAME)
            elements.append(Paragraph("身份統計 (按不重複車輛數量)", identity_title_style))
            elements.append(Spacer(1, 10))

            identity_data = [["身份", "入場車輛", "無法入場車輛", "預警車輛", "合計"]]
            all_identities = set(list(stats.get('no_entry_identities', {}).keys()) + 
                               list(stats.get('entry_identities', {}).keys()) +
                               list(stats.get('warning_identities', {}).keys()))
            for identity in sorted(all_identities):
                no_entry_count = stats.get('no_entry_identities', {}).get(identity, 0)
                entry_count = stats.get('entry_identities', {}).get(identity, 0)
                warning_count = stats.get('warning_identities', {}).get(identity, 0)
                total = no_entry_count + entry_count + warning_count
                identity_data.append([identity, str(entry_count), str(no_entry_count), str(warning_count), str(total)])

            identity_table = Table(identity_data, colWidths=[1.0*inch, 1.0*inch, 1.2*inch, 1.0*inch, 1.0*inch])
            identity_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(identity_table)

            # 生成 PDF
            doc.build(elements)
            return True, "PDF 報表已生成: {0}".format(output_path)

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            return False, "生成 PDF 失敗:\n{0}\n詳細: {1}".format(str(e), error_detail)

    @staticmethod
    def generate_details_report(output_path: str, records: List[Dict], title: str = "Vehicle Traffic Details") -> Tuple[bool, str]:
        """生成詳細記錄 PDF"""
        if not REPORTLAB_AVAILABLE:
            return False, "缺少 reportlab 庫\n請執行: pip install reportlab"

        try:
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    return False, "無法創建輸出目錄: {0}".format(str(e))
            
            test_path = output_path
            if not output_dir:
                test_path = os.path.join(os.getcwd(), output_path)
            
            try:
                test_dir = os.path.dirname(test_path) or os.getcwd()
                if not os.access(test_dir, os.W_OK):
                    return False, "無寫入權限到目錄: {0}".format(test_dir)
            except Exception as e:
                return False, "權限檢查失敗: {0}".format(str(e))
            
            doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()

            # 標題
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=12,
                alignment=TA_CENTER,
                fontName=FONT_NAME
            )
            title_text = "車輛通行詳細報表"
            elements.append(Paragraph(title_text, title_style))
            elements.append(Spacer(1, 12))

            # 生成時間和記錄數
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT, fontName=FONT_NAME)
            elements.append(Paragraph("生成時間: {0} | 總共: {1} 筆記錄".format(now, len(records)), info_style))
            elements.append(Spacer(1, 12))

            # 詳細表格（分頁）
            page_size = 50
            total_pages = (len(records) + page_size - 1) // page_size
            
            for page_num in range(total_pages):
                if page_num > 0:
                    elements.append(PageBreak())
                    elements.append(Paragraph("第 {0}/{1} 頁".format(page_num + 1, total_pages), info_style))
                    elements.append(Spacer(1, 10))
                
                start_idx = page_num * page_size
                end_idx = min(start_idx + page_size, len(records))
                page_records = records[start_idx:end_idx]
                
                data = [["ID", "時間", "車道", "車號", "身份", "結果", "標籤"]]
                for record in page_records:
                    data.append([
                        str(record['id']),
                        record['time'],
                        record['lane'],
                        record['plate_number'],
                        record['identity'],
                        record['result'],
                        record['etag']
                    ])

                table = Table(data, colWidths=[0.6*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                elements.append(table)

            # 生成 PDF
            doc.build(elements)
            return True, "PDF 報表已生成: {0}".format(output_path)

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            return False, "生成 PDF 失敗:\n{0}\n詳細: {1}".format(str(e), error_detail)


# ============================================================================
# 媒體管理器
# ============================================================================

class MediaManager:
    """照片和視頻管理"""

    _media_cache = {}
    FIXED_WIDTH = 380
    FIXED_HEIGHT = 280

    @staticmethod
    def load_and_display_photo(label: tk.Label, photo_path: str) -> bool:
        """加載並顯示照片"""
        try:
            if not os.path.exists(photo_path):
                label.config(text="✘ 照片文件不存在", fg="#f44336", image="")
                return False

            img = Image.open(photo_path)

            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', (img.width, img.height), (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                if img.mode == 'RGBA' or img.mode == 'LA':
                    background.paste(img, mask=img.split()[-1])
                    img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            img_width, img_height = img.size
            scale = min(MediaManager.FIXED_WIDTH / img_width, MediaManager.FIXED_HEIGHT / img_height)
            new_width = int(img_width * scale)
            new_height = int(img_height * scale)

            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

            background = Image.new('RGB', (MediaManager.FIXED_WIDTH, MediaManager.FIXED_HEIGHT), (240, 240, 240))
            left = (MediaManager.FIXED_WIDTH - new_width) // 2
            top = (MediaManager.FIXED_HEIGHT - new_height) // 2
            background.paste(img, (left, top))

            photo = ImageTk.PhotoImage(background)
            label.config(image=photo, text="", fg="white")
            MediaManager._media_cache[id(label)] = photo
            return True
        except:
            label.config(text="✘ 無法加載照片", fg="#f44336", image="")
            return False

    @staticmethod
    def check_ffmpeg() -> bool:
        """檢查 ffmpeg 是否可用"""
        try:
            cmd = ['ffmpeg', '-version'] if platform.system() != 'Windows' else ['ffmpeg', '-version']
            result = subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=5)
            return result.returncode == 0
        except:
            return False

    @staticmethod
    def extract_video_frame(video_path: str) -> str:
        """使用 ffmpeg 提取視頻第一幀"""
        try:
            if not os.path.exists(video_path):
                return None
            if not MediaManager.check_ffmpeg():
                return None

            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
                temp_frame_path = tmp_file.name

            cmd = ['ffmpeg', '-i', video_path, '-vframes', '1', '-q:v', '2', '-y', temp_frame_path]

            result = subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=30)

            if result.returncode != 0 or not os.path.exists(temp_frame_path):
                return None

            return temp_frame_path
        except:
            return None

    @staticmethod
    def load_and_display_video(label: tk.Label, video_path: str, play_callback=None) -> bool:
        """加載並顯示視頻第一幀"""
        try:
            if not os.path.exists(video_path):
                label.config(text="✘ 影像文件不存在\n點擊播放", fg="#f44336", image="")
                return False

            frame_path = MediaManager.extract_video_frame(video_path)

            if not frame_path:
                label.config(text="✘ 無法提取影像幀\n請確保已安裝 ffmpeg", fg="#f44336", image="")
                return False

            img = Image.open(frame_path)

            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', (img.width, img.height), (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                if img.mode == 'RGBA' or img.mode == 'LA':
                    background.paste(img, mask=img.split()[-1])
                    img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            img_width, img_height = img.size
            scale = min(MediaManager.FIXED_WIDTH / img_width, MediaManager.FIXED_HEIGHT / img_height)
            new_width = int(img_width * scale)
            new_height = int(img_height * scale)

            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

            background = Image.new('RGB', (MediaManager.FIXED_WIDTH, MediaManager.FIXED_HEIGHT), (240, 240, 240))
            left = (MediaManager.FIXED_WIDTH - new_width) // 2
            top = (MediaManager.FIXED_HEIGHT - new_height) // 2
            background.paste(img, (left, top))

            photo = ImageTk.PhotoImage(background)
            label.config(image=photo, text="🎬 點擊播放", fg="#2196F3")
            MediaManager._media_cache[id(label)] = photo

            if play_callback:
                label.bind("<Button-1>", lambda e: play_callback())

            try:
                os.remove(frame_path)
            except:
                pass

            return True
        except:
            label.config(text="✘ 無法加載影像", fg="#f44336", image="")
            return False


# ============================================================================
# 統計窗口
# ============================================================================

class StatisticsWindow(tk.Toplevel):
    """統計資訊窗口"""

    def __init__(self, parent, stats: Dict[str, Any], importer: VehicleImporter):
        super().__init__(parent)
        self.title("📊 統計資訊")
        self.geometry("1400x800")
        self.resizable(True, True)
        self.configure(bg="white")
        self.transient(parent)

        self.stats = stats
        self.importer = importer
        self.parent = parent

        self._build_ui()

    def _build_ui(self):
        """構建統計 UI"""
        title = tk.Label(
            self,
            text="📊 統計資訊 (按不重複車輛數量統計)",
            font=("微軟雅黑", 16, "bold"),
            bg="white",
            fg="#333"
        )
        title.pack(pady=15)

        info_frame = tk.Frame(self, bg="#f5f5f5", relief=tk.SUNKEN, bd=2)
        info_frame.pack(fill=tk.X, padx=15, pady=(0, 15))

        no_entry_percent = round(self.stats['no_entry_total'] * 100 / self.stats['total']) if self.stats['total'] > 0 else 0
        entry_percent = round(self.stats['entry_total'] * 100 / self.stats['total']) if self.stats['total'] > 0 else 0
        warning_percent = round(self.stats['warning_total'] * 100 / self.stats['total']) if self.stats['total'] > 0 else 0

        info_text = "📋 總計: {0} 輛車  |  ✘ 無法入場: {1} 輛 ({2}%)  |  ✓ 入場: {3} 輛 ({4}%)  |  ⚠️ 預警: {5} 輛 ({6}%)".format(
            self.stats['total'],
            self.stats['no_entry_total'],
            no_entry_percent,
            self.stats['entry_total'],
            entry_percent,
            self.stats['warning_total'],
            warning_percent
        )
        
        if self.stats['date_range'][0]:
            info_text += "  |  📅 日期: {0} ~ {1}".format(
                self.stats['date_range'][0],
                self.stats['date_range'][1]
            )

        tk.Label(
            info_frame,
            text=info_text,
            font=("微軟雅黑", 11, "bold"),
            bg="#f5f5f5",
            fg="#333",
            anchor=tk.W,
            justify=tk.LEFT
        ).pack(fill=tk.X, padx=10, pady=10)

        # 選項卡
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 選項卡 1: 無法入場車子
        no_entry_frame = tk.Frame(notebook)
        notebook.add(no_entry_frame, text="❌ 無法入場車子")
        self._build_no_entry_tab(no_entry_frame)

        # 選項卡 2: 身份統計
        identity_frame = tk.Frame(notebook)
        notebook.add(identity_frame, text="👥 身份統計")
        self._build_identity_tab(identity_frame)

        # 按鈕欄
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 16))

        tk.Button(btn_frame, text="📊 導出統計PDF", command=self.export_statistics_pdf,
                  bg="#2196F3", fg="white", activebackground="#1976D2",
                  font=("微軟雅黑", 10), padx=20, pady=6,
                  relief="flat", bd=0, cursor="hand2").pack(side=tk.LEFT, padx=4)

        tk.Button(btn_frame, text="📋 導出詳細PDF", command=self.export_details_pdf,
                  bg="#4CAF50", fg="white", activebackground="#388E3C",
                  font=("微軟雅黑", 10), padx=20, pady=6,
                  relief="flat", bd=0, cursor="hand2").pack(side=tk.LEFT, padx=4)

        tk.Button(btn_frame, text="✘ 關閉", command=self.destroy,
                  bg="#757575", fg="white", activebackground="#666",
                  font=("微軟雅黑", 10), padx=20, pady=6,
                  relief="flat", bd=0, cursor="hand2").pack(side=tk.LEFT, padx=4)

    def _build_no_entry_tab(self, parent):
        """構建無法入場選項卡"""
        scroll_frame = tk.Frame(parent)
        scroll_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar_y = ttk.Scrollbar(scroll_frame, orient="vertical")
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        scrollbar_x = ttk.Scrollbar(scroll_frame, orient="horizontal")
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        columns = ("車號", "無法入場天數", "最後無法入場時間")
        tree = ttk.Treeview(scroll_frame, columns=columns, height=25, show="headings", yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        tree.column("車號", width=150, anchor="center")
        tree.column("無法入場天數", width=150, anchor="center")
        tree.column("最後無法入場時間", width=250, anchor="center")

        tree.heading("車號", text="車號")
        tree.heading("無法入場天數", text="無法入場天數")
        tree.heading("最後無法入場時間", text="最後無法入場時間")

        tree.tag_configure("data", background="#f5f5f5")

        for plate, count in sorted(self.stats['no_entry_vehicles'].items(), key=lambda x: x[1], reverse=True):
            last_time = self.stats.get('no_entry_last_time', {}).get(plate, "")
            tree.insert("", tk.END, values=(plate, count, last_time), tags=("data",))

        tree.pack(fill=tk.BOTH, expand=True)

    def _build_identity_tab(self, parent):
        """構建身份統計選項卡 - ✅ 按不重複車輛數量"""
        scroll_frame = tk.Frame(parent)
        scroll_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar_y = ttk.Scrollbar(scroll_frame, orient="vertical")
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        scrollbar_x = ttk.Scrollbar(scroll_frame, orient="horizontal")
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        columns = ("身份", "入場車輛", "無法入場車輛", "預警車輛", "合計")
        tree = ttk.Treeview(scroll_frame, columns=columns, height=25, show="headings", yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        tree.column("身份", width=100, anchor="center")
        tree.column("入場車輛", width=120, anchor="center")
        tree.column("無法入場車輛", width=150, anchor="center")
        tree.column("預警車輛", width=120, anchor="center")
        tree.column("合計", width=100, anchor="center")

        tree.heading("身份", text="身份")
        tree.heading("入場車輛", text="✓ 入場車輛")
        tree.heading("無法入場車輛", text="✘ 無法入場車輛")
        tree.heading("預警車輛", text="⚠️ 預警車輛")
        tree.heading("合計", text="合計")

        tree.tag_configure("data", background="#f5f5f5")

        all_identities = set(list(self.stats.get('no_entry_identities', {}).keys()) + 
                           list(self.stats.get('entry_identities', {}).keys()) +
                           list(self.stats.get('warning_identities', {}).keys()))
        for identity in sorted(all_identities):
            no_entry_count = self.stats.get('no_entry_identities', {}).get(identity, 0)
            entry_count = self.stats.get('entry_identities', {}).get(identity, 0)
            warning_count = self.stats.get('warning_identities', {}).get(identity, 0)
            total = no_entry_count + entry_count + warning_count
            tree.insert("", tk.END, values=(identity, entry_count, no_entry_count, warning_count, total), tags=("data",))

        tree.pack(fill=tk.BOTH, expand=True)

    def export_statistics_pdf(self):
        """導出統計 PDF"""
        file_path = filedialog.asksaveasfilename(
            title="保存統計報表",
            defaultextension=".pdf",
            filetypes=[("PDF 文件", "*.pdf")]
        )
        if not file_path:
            return

        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("錯誤", "缺少 reportlab 庫\n請執行: pip install reportlab")
            return

        try:
            test_dir = os.path.dirname(file_path)
            if test_dir and not os.path.exists(test_dir):
                os.makedirs(test_dir)
            if not os.access(test_dir if test_dir else '.', os.W_OK):
                messagebox.showerror("錯誤", "無寫入權限到: {0}".format(test_dir if test_dir else '當前目錄'))
                return
        except Exception as e:
            messagebox.showerror("錯誤", "路徑檢查失敗: {0}".format(str(e)))
            return

        success, message = PDFReportGenerator.generate_statistics_report(file_path, self.stats)
        if success:
            messagebox.showinfo("成功", message)
        else:
            messagebox.showerror("失敗", message)

    def export_details_pdf(self):
        """導出詳細 PDF"""
        file_path = filedialog.asksaveasfilename(
            title="保存詳細報表",
            defaultextension=".pdf",
            filetypes=[("PDF 文件", "*.pdf")]
        )
        if not file_path:
            return

        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("錯誤", "缺少 reportlab 庫\n請執行: pip install reportlab")
            return

        try:
            test_dir = os.path.dirname(file_path)
            if test_dir and not os.path.exists(test_dir):
                os.makedirs(test_dir)
            if not os.access(test_dir if test_dir else '.', os.W_OK):
                messagebox.showerror("錯誤", "無寫入權限到: {0}".format(test_dir if test_dir else '當前目錄'))
                return
        except Exception as e:
            messagebox.showerror("錯誤", "路徑檢查失敗: {0}".format(str(e)))
            return

        records = self.importer.get_all_records()
        success, message = PDFReportGenerator.generate_details_report(file_path, records)
        if success:
            messagebox.showinfo("成功", message)
        else:
            messagebox.showerror("失敗", message)


# ============================================================================
# GUI 層 - 其餘代碼保持不變
# ============================================================================

class VehicleImporterGUI:
    """主 GUI 應用"""

    def __init__(self, root):
        self.root = root
        self.importer = VehicleImporter()

        if isinstance(self.root, (tk.Tk, tk.Toplevel)):
            self.root.title("南方莊園車輛通行記錄導入管理系統")
            self.root.geometry("1600x900")
            self.root.resizable(True, True)
            self.root.configure(bg="#f0f0f0")

        self.progress_window = None
        self.current_video_path = None

        self.create_menu()
        self.create_main_ui()
        self.refresh_table()

    def create_menu(self):
        """創建菜單"""
        if not isinstance(self.root, (tk.Tk, tk.Toplevel)):
            return

        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="檔案", menu=file_menu)
        file_menu.add_command(label="📁 設置照片目錄", command=self.set_photo_directory)
        file_menu.add_command(label="🎬 設置影像目錄", command=self.set_vehicle_image_directory)
        file_menu.add_separator()
        file_menu.add_command(label="📥 追加匯入", command=self.import_append)
        file_menu.add_command(label="🔄 覆蓋匯入", command=self.import_replace)
        file_menu.add_separator()
        file_menu.add_command(label="📊 導出統計PDF", command=self.export_statistics_pdf)
        file_menu.add_command(label="📋 導出詳細PDF", command=self.export_details_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="🚪 結束", command=self.root.quit)

        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="查看", menu=view_menu)
        view_menu.add_command(label="🔄 刷新", command=self.refresh_table)
        view_menu.add_command(label="📊 統計", command=self.show_statistics)

    def set_photo_directory(self):
        """設置照片目錄"""
        directory = filedialog.askdirectory(title="選擇車輛照片目錄")
        if directory:
            self.importer.set_photo_directory(directory)
            messagebox.showinfo("成功", "照片目錄已設置")

    def set_vehicle_image_directory(self):
        """設置影像目錄"""
        directory = filedialog.askdirectory(title="選擇車輛影像目錄")
        if directory:
            self.importer.set_vehicle_image_directory(directory)
            messagebox.showinfo("成功", "影像目錄已設置")

    def create_main_ui(self):
        """創建主 UI"""
        top_frame = tk.Frame(self.root, bg="#f0f0f0")
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(top_frame, text="🚗 車輛通行記錄管理", font=("微軟雅黑", 20, "bold"), bg="#f0f0f0", fg="#333").pack(side=tk.LEFT)

        button_frame = tk.Frame(top_frame, bg="#f0f0f0")
        button_frame.pack(side=tk.RIGHT)

        buttons = [
            ("📁 照片", self.set_photo_directory, "#FF5722"),
            ("🎬 影像", self.set_vehicle_image_directory, "#00A8E8"),
            ("📥 追加", self.import_append, "#4CAF50"),
            ("🔄 覆蓋", self.import_replace, "#FF9800"),
            ("📊 統計", self.show_statistics, "#9C27B0"),
            ("📋 報表", self.export_statistics_pdf, "#F57C00"),
        ]

        for text, command, color in buttons:
            tk.Button(button_frame, text=text, command=command, bg=color, fg="white", padx=12, pady=8, font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.LEFT, padx=3)

        ttk.Separator(self.root, orient="horizontal").pack(fill=tk.X)

        search_frame = tk.LabelFrame(self.root, text="🔍 查詢條件", font=("微軟雅黑", 10, "bold"), bg="#f5f5f5", padx=10, pady=8)
        search_frame.pack(fill=tk.X, padx=10, pady=8)

        unique_values = self.importer.get_unique_values()

        row1 = tk.Frame(search_frame, bg="#f5f5f5")
        row1.pack(fill=tk.X, pady=3)

        tk.Label(row1, text="車號:", font=("微軟雅黑", 9), bg="#f5f5f5", width=6).pack(side=tk.LEFT)
        self.plate_var = tk.StringVar()
        tk.Entry(row1, textvariable=self.plate_var, width=20, font=("微軟雅黑", 9)).pack(side=tk.LEFT, padx=(0, 10))

        tk.Label(row1, text="身份:", font=("微軟雅黑", 9), bg="#f5f5f5", width=6).pack(side=tk.LEFT)
        self.identity_var = tk.StringVar()
        ttk.Combobox(row1, textvariable=self.identity_var, values=[""] + unique_values["identities"], width=15, state="readonly").pack(side=tk.LEFT, padx=(0, 10))

        tk.Label(row1, text="結果:", font=("微軟雅黑", 9), bg="#f5f5f5", width=6).pack(side=tk.LEFT)
        self.result_var = tk.StringVar()
        ttk.Combobox(row1, textvariable=self.result_var, values=[""] + unique_values["results"], width=15, state="readonly").pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(row1, text="🔍 查詢", command=self.search_records, bg="#00BCD4", fg="white", padx=15, font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.LEFT, padx=2)
        tk.Button(row1, text="清空", command=self.clear_search, bg="#757575", fg="white", padx=15, font=("微軟雅黑", 9), cursor="hand2").pack(side=tk.LEFT, padx=2)

        main_container = tk.Frame(self.root, bg="white")
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        left_frame = tk.Frame(main_container, bg="white")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        self.status_label = tk.Label(left_frame, text="", font=("微軟雅黑", 9), bg="white", fg="#666")
        self.status_label.pack(anchor=tk.W, pady=(0, 3))

        scroll_frame = tk.Frame(left_frame, bg="white")
        scroll_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar_y = ttk.Scrollbar(scroll_frame, orient="vertical")
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        scrollbar_x = ttk.Scrollbar(scroll_frame, orient="horizontal")
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        columns = ("ID", "時間", "車號", "身份", "結果", "標籤")
        self.tree = ttk.Treeview(scroll_frame, columns=columns, height=20, show="headings", yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)

        for col in columns:
            self.tree.column(col, width=100, anchor="center")
            self.tree.heading(col, text=col)

        self.tree.tag_configure("entry", background="#d4edda")
        self.tree.tag_configure("no_entry", background="#f8d7da")
        self.tree.tag_configure("warning", background="#fff3cd")

        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_record_selected)

        right_frame = tk.Frame(main_container, bg="white")
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        photo_frame = tk.LabelFrame(right_frame, text="📷 車輛照片 (380x280)", font=("微軟雅黑", 10, "bold"), bg="white")
        photo_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 3))

        photo_container = tk.Frame(photo_frame, bg="#f0f0f0", width=380, height=280)
        photo_container.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        photo_container.pack_propagate(False)

        self.photo_display = tk.Label(photo_container, text="選擇記錄查看照片", font=("微軟雅黑", 10), bg="#e8e8e8", fg="#999", relief=tk.SUNKEN, bd=2)
        self.photo_display.pack(fill=tk.BOTH, expand=True)

        video_frame = tk.LabelFrame(right_frame, text="🎬 車輛影片 (380x280)", font=("微軟雅黑", 10, "bold"), bg="white")
        video_frame.pack(fill=tk.BOTH, expand=True, pady=(3, 0))

        video_container = tk.Frame(video_frame, bg="#f0f0f0", width=380, height=280)
        video_container.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        video_container.pack_propagate(False)

        self.video_display = tk.Label(video_container, text="選擇記錄顯示影片", font=("微軟雅黑", 10), bg="#e8e8e8", fg="#999", relief=tk.SUNKEN, bd=2, cursor="hand2")
        self.video_display.pack(fill=tk.BOTH, expand=True)

    def on_record_selected(self, event):
        """選擇記錄時顯示照片和影像"""
        selection = self.tree.selection()
        if not selection:
            return

        item_id = selection[0]
        values = self.tree.item(item_id)['values']

        time_str = values[1]
        plate_number = values[2]

        if not plate_number:
            self.photo_display.config(text="⚠️ 無車號", fg="#f44336", image="")
            self.video_display.config(text="⚠️ 無車號", fg="#f44336", image="")
            return

        photo_path = self.importer.find_vehicle_photo(plate_number, time_str)
        if photo_path:
            def load_photo():
                MediaManager.load_and_display_photo(self.photo_display, photo_path)
            threading.Thread(target=load_photo, daemon=True).start()
        else:
            self.photo_display.config(text="📷 照片未找到", fg="#f44336", image="")

        image_path = self.importer.find_vehicle_image(plate_number, time_str)
        if image_path and os.path.exists(image_path):
            self.current_video_path = image_path
            def load_video():
                MediaManager.load_and_display_video(self.video_display, image_path, play_callback=lambda: self.play_video(image_path))
            threading.Thread(target=load_video, daemon=True).start()
        else:
            self.video_display.config(text="🎬 影片未找到", fg="#f44336", image="")

    def play_video(self, video_path: str):
        """播放影片"""
        try:
            if os.name == 'nt':
                os.startfile(video_path)
            elif platform.system() == 'Darwin':
                subprocess.Popen(['open', video_path])
            else:
                subprocess.Popen(['xdg-open', video_path])
        except Exception as e:
            messagebox.showerror("錯誤", "無法播放影片: {0}".format(str(e)))

    def import_append(self):
        """追加匯入"""
        file_path = filedialog.askopenfilename(title="選擇檔案", filetypes=[("所有格式", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")])
        if file_path:
            self.perform_import(file_path, append_mode=True)

    def import_replace(self):
        """覆蓋匯入"""
        file_path = filedialog.askopenfilename(title="選擇檔案", filetypes=[("所有格式", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")])
        if file_path and messagebox.askyesno("確認", "確定要覆蓋現有數據嗎？"):
            self.perform_import(file_path, append_mode=False)

    def perform_import(self, file_path: str, append_mode: bool):
        """執行匯入"""
        self.progress_window = ProgressWindow(self.root, "📥 匯入進度")
        self.importer.set_progress_callback(self.progress_window.update_progress)

        def import_thread():
            try:
                success, message, count = self.importer.import_from_file(file_path, append_mode=append_mode)
                self.root.after(500, lambda: self.progress_window.close())
                if success:
                    self.root.after(600, lambda: messagebox.showinfo("成功", message))
                    self.root.after(600, lambda: self.refresh_table())
                else:
                    self.root.after(600, lambda: messagebox.showerror("失敗", message))
            except Exception as e:
                self.root.after(500, lambda: self.progress_window.close())
                self.root.after(600, lambda: messagebox.showerror("錯誤", str(e)))

        threading.Thread(target=import_thread, daemon=True).start()

    def refresh_table(self):
        """刷新表格"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        records = self.importer.get_all_records()

        for record in records:
            tag = "entry" if record['result'] == "入場" else "no_entry" if record['result'] == "無法入場" else "warning" if record['result'] == "預警" else ""
            self.tree.insert("", tk.END, values=(record['id'], record['time'], record['plate_number'], record['identity'], record['result'], record['etag']), tags=(tag,))

        self.status_label.config(text="總共 {0} 筆記錄".format(len(records)))

    def search_records(self):
        """查詢記錄"""
        plate_number = self.plate_var.get().strip()
        identity = self.identity_var.get().strip()
        result = self.result_var.get().strip()

        if not plate_number and not identity and not result:
            messagebox.showwarning("警告", "請至少選擇一個查詢條件")
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        search_results = self.importer.search_by_conditions(plate_number=plate_number, identity=identity, result=result)

        if not search_results:
            messagebox.showinfo("查詢結果", "未找到匹配的記錄")
            self.status_label.config(text="查詢結果: 0 筆記錄")
            return

        for record in search_results:
            tag = "entry" if record['result'] == "入場" else "no_entry" if record['result'] == "無法入場" else "warning" if record['result'] == "預警" else ""
            self.tree.insert("", tk.END, values=(record['id'], record['time'], record['plate_number'], record['identity'], record['result'], record['etag']), tags=(tag,))

        self.status_label.config(text="查詢結果: {0} 筆記錄".format(len(search_results)))

    def clear_search(self):
        """清空搜索"""
        self.plate_var.set("")
        self.identity_var.set("")
        self.result_var.set("")
        self.refresh_table()

    def show_statistics(self):
        """顯示統計"""
        stats = self.importer.get_statistics()

        if stats['total'] == 0:
            messagebox.showinfo("統計資訊", "暫無數據")
            return

        stat_window = StatisticsWindow(self.root, stats, self.importer)

    def export_statistics_pdf(self):
        """導出統計 PDF"""
        file_path = filedialog.asksaveasfilename(
            title="保存統計報表",
            defaultextension=".pdf",
            filetypes=[("PDF 文件", "*.pdf")]
        )
        if not file_path:
            return

        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("錯誤", "缺少 reportlab 庫\n請執行: pip install reportlab")
            return

        try:
            test_dir = os.path.dirname(file_path)
            if test_dir and not os.path.exists(test_dir):
                os.makedirs(test_dir)
            if not os.access(test_dir if test_dir else '.', os.W_OK):
                messagebox.showerror("錯誤", "無寫入權限到: {0}".format(test_dir if test_dir else '當前目錄'))
                return
        except Exception as e:
            messagebox.showerror("錯誤", "路徑檢查失敗: {0}".format(str(e)))
            return

        stats = self.importer.get_statistics()
        if stats['total'] == 0:
            messagebox.showwarning("警告", "暫無數據無法生成報表")
            return

        success, message = PDFReportGenerator.generate_statistics_report(file_path, stats)
        if success:
            messagebox.showinfo("成功", message)
        else:
            messagebox.showerror("失敗", message)

    def export_details_pdf(self):
        """導出詳細 PDF"""
        file_path = filedialog.asksaveasfilename(
            title="保存詳細報表",
            defaultextension=".pdf",
            filetypes=[("PDF 文件", "*.pdf")]
        )
        if not file_path:
            return

        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("錯誤", "缺少 reportlab 庫\n請執行: pip install reportlab")
            return

        try:
            test_dir = os.path.dirname(file_path)
            if test_dir and not os.path.exists(test_dir):
                os.makedirs(test_dir)
            if not os.access(test_dir if test_dir else '.', os.W_OK):
                messagebox.showerror("錯誤", "無寫入權限到: {0}".format(test_dir if test_dir else '當前目錄'))
                return
        except Exception as e:
            messagebox.showerror("錯誤", "路徑檢查失敗: {0}".format(str(e)))
            return

        records = self.importer.get_all_records()
        if not records:
            messagebox.showwarning("警告", "暫無數據無法生成報表")
            return

        success, message = PDFReportGenerator.generate_details_report(file_path, records)
        if success:
            messagebox.showinfo("成功", message)
        else:
            messagebox.showerror("失敗", message)


# ============================================================================
# 嵌入式視圖
# ============================================================================

class VehicleImporterView(ttk.Frame):
    """給 launcher 嵌入的 Qphoto 頁面"""
    def __init__(self, parent):
        super().__init__(parent)
        host = tk.Frame(self, bg="#f0f0f0")
        host.pack(fill=tk.BOTH, expand=True)
        self.app = VehicleImporterGUI(host)


# ============================================================================
# 主程序
# ============================================================================

def main():
    """主程序入口"""
    root = tk.Tk()
    app = VehicleImporterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
